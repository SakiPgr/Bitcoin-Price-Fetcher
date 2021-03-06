import requests
import datetime
import accesskey  # I import url+accesskey to get conversion rate to euro
import os.path as path
from openpyxl import Workbook, load_workbook


BITCOIN_PRICE_URL = 'https://api.coinmarketcap.com/v1/ticker/bitcoin/'
excel_file = 'bitcoinprice.xlsx'
column_names = ["Ημερομηνία",
                "Ώρα",
                "Τιμή σε Δολάριο",
                "Τιμή σε Ευρώ",
                "Τιμή Δολαρίου"]


def main():
    CheckForWorkbook(excel_file).save(excel_file)


def CheckForWorkbook(book):
    """ Creates the file on first use and after every consecutive
    use adds a new line with the information. """
    if path.isfile(book):
        wb = load_workbook(book)
        return ConsecutiveEntry(wb)
    else:
        return FirstEntry()


def FirstEntry():
    wb = Workbook()
    ws = wb.active
    ws.append(column_names)
    ConsecutiveEntry(wb)
    return wb


def ConsecutiveEntry(book):
    ws = book.active
    ws.append([DateAndTime()[0],
               DateAndTime()[1],
               RoundNumbers(bitcoin_dollar),
               RoundNumbers(BitcoinPriceInEuro()),
               GetDollarToEuroConversionRate(accesskey.AccessKey())])
    return book


def DateAndTime():
    date_now = datetime.datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.datetime.now().strftime('%H:%M')
    return (date_now, time_now)


#  The price in Dollar of Bitcoin
def GetBitcoinPrice(url):
    response = requests.get(BITCOIN_PRICE_URL)
    response_json = response.json()
    bitcoin_price_str = response_json[0]['price_usd']
    return bitcoin_price_str


def BitcoinPriceInEuro():
    price_in_euro = bitcoin_dollar * dollar_to_euro
    return price_in_euro


def RoundNumbers(num):
    return round(float(num), 2)


# The Dollar to Euro conversion rate
def GetDollarToEuroConversionRate(url):
    response = requests.get(url)
    response_json = response.json()
    return float(response_json['quotes']['USDEUR'])


bitcoin_dollar = float(GetBitcoinPrice(BITCOIN_PRICE_URL))
# I use an API to get the dollar to euro conversion with a key bound to my
# email. The page I use is https://apilayer.com/.
dollar_to_euro = GetDollarToEuroConversionRate(accesskey.AccessKey())

if __name__ == "__main__":
    main()
